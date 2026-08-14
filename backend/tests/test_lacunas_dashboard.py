"""Cobertura das lacunas de back-end levantadas em `resume/06`.

Mesmo padrão das outras suítes: função pura, contrato de schema e registro de
rota. O que depende de SQL — `date_trunc` no fuso local, `DISTINCT` de
categorias, ordenação estável — não é alcançável aqui e está listado como
validação manual em `resume/07-lacunas-backend-implementadas.md`.
"""

from datetime import UTC, date, datetime, timedelta
from decimal import Decimal
from types import SimpleNamespace
import io
import json
import uuid

from fastapi import HTTPException
from fastapi.security import HTTPAuthorizationCredentials
from openpyxl import load_workbook
from pydantic import ValidationError
import pytest

from app.api.dependencies import get_current_user, get_optional_user
from app.api.v1 import auth
from app.api.v1.transactions import ORDERABLE_COLUMNS, OrderBy
from app.core.security import generate_refresh_token, hash_refresh_token
from app.main import app
from app.models.refresh_token import RefreshToken
from app.schemas.auth import (
    ChangePasswordRequest,
    LoginRequest,
    LogoutRequest,
    ProfileUpdate,
    TokenResponse,
)
from app.services.login_throttle import (
    FAILURES_PER_LOCK,
    FAILURE_WINDOW,
    IP_TOLERANCE,
    LEVEL_DECAY,
    LOCK_DURATIONS,
    LoginBlocked,
    ThrottleState,
    forget_stale_failures,
    lock_duration,
    register_failure,
    scope_hash,
    seconds_until_release,
)
from app.schemas.investment import MovementCreate, MovementUpdate
from app.schemas.transaction import CategoryOption, TimeseriesPoint
from app.services.sessions import (
    RefreshTokenError,
    revoke_refresh_token,
    rotate_refresh_token,
)
from app.services.spreadsheets import export_portfolio_xlsx, export_positions_csv
from app.services.timeseries import (
    MAX_POINTS,
    SeriesPoint,
    TimeseriesRangeError,
    bucket_range,
    bucket_start,
    build_series,
    next_bucket,
)


# ---------------------------------------------------------------------------
# A1 — série temporal
# ---------------------------------------------------------------------------


def test_week_buckets_start_on_monday_like_postgres() -> None:
    # 12/08/2026 é uma quarta-feira; a segunda daquela semana é dia 10.
    assert bucket_start(date(2026, 8, 12), "week") == date(2026, 8, 10)
    assert bucket_start(date(2026, 8, 10), "week") == date(2026, 8, 10)
    assert bucket_start(date(2026, 8, 12), "month") == date(2026, 8, 1)
    assert bucket_start(date(2026, 8, 12), "day") == date(2026, 8, 12)


def test_month_bucket_crosses_the_year() -> None:
    assert next_bucket(date(2026, 12, 1), "month") == date(2027, 1, 1)
    assert next_bucket(date(2026, 1, 31), "day") == date(2026, 2, 1)
    assert next_bucket(date(2026, 8, 10), "week") == date(2026, 8, 17)


def test_series_fills_periods_without_transactions() -> None:
    totals = {
        (date(2026, 6, 1), "income"): Decimal("5000.00"),
        (date(2026, 8, 1), "expense"): Decimal("300.00"),
    }
    points = build_series(totals, date(2026, 6, 15), date(2026, 8, 20), "month")

    assert [point.period for point in points] == [
        date(2026, 6, 1),
        date(2026, 7, 1),
        date(2026, 8, 1),
    ]
    # Julho não teve lançamento: precisa aparecer como zero, não sumir do eixo.
    assert points[1] == SeriesPoint(
        period=date(2026, 7, 1),
        income=Decimal("0.00"),
        expense=Decimal("0.00"),
        balance=Decimal("0.00"),
    )
    assert points[0].balance == Decimal("5000.00")
    assert points[2].balance == Decimal("-300.00")


def test_series_refuses_an_interval_that_would_explode_the_payload() -> None:
    first = date(2020, 1, 1)
    last = first + timedelta(days=MAX_POINTS + 10)
    with pytest.raises(TimeseriesRangeError, match=str(MAX_POINTS)):
        bucket_range(first, last, "day")
    # A mesma janela em meses é pequena e continua permitida.
    assert len(bucket_range(first, last, "month")) < MAX_POINTS


def test_timeseries_point_serializes_money_as_string() -> None:
    point = TimeseriesPoint(
        period=date(2026, 8, 1),
        income=Decimal("10.50"),
        expense=Decimal("0.00"),
        balance=Decimal("10.50"),
    )
    payload = json.loads(point.model_dump_json())
    assert payload == {
        "period": "2026-08-01",
        "income": "10.50",
        "expense": "0.00",
        "balance": "10.50",
    }


# ---------------------------------------------------------------------------
# A3 / A4 — categorias e ordenação
# ---------------------------------------------------------------------------


def test_category_option_carries_usage_count() -> None:
    option = CategoryOption(category="Alimentação", count=12)
    assert json.loads(option.model_dump_json()) == {
        "category": "Alimentação",
        "count": 12,
    }


def test_order_by_whitelist_matches_the_declared_literal() -> None:
    """A lista branca e o `Literal` da query precisam andar juntos.

    Se divergissem, ou uma coluna aceita pelo tipo faltaria no dicionário
    (KeyError em produção), ou o dicionário exporia uma coluna que a API não
    deveria deixar ordenar.
    """
    assert set(OrderBy.__args__) == set(ORDERABLE_COLUMNS)


# ---------------------------------------------------------------------------
# C1 / A8 — sessões e senha
# ---------------------------------------------------------------------------


def test_refresh_token_is_opaque_and_stored_only_as_a_hash() -> None:
    token, token_hash, expires_at = generate_refresh_token()

    assert expires_at > datetime.now(UTC)
    assert hash_refresh_token(token) == token_hash
    assert len(token_hash) == 64  # SHA-256 em hexadecimal cabe no String(64).
    # O valor entregue ao cliente não pode ser derivável do que fica no banco.
    assert token not in token_hash

    outro, _, _ = generate_refresh_token()
    assert outro != token


def test_token_response_requires_the_refresh_pair() -> None:
    """O front precisa receber os dois tokens; sem isso a sessão não renova."""
    with pytest.raises(ValidationError, match="refresh_token"):
        TokenResponse(
            access_token="abc",
            expires_at=datetime.now(UTC),
            user={
                "id": uuid.uuid4(),
                "email": "pessoa@example.com",
                "full_name": None,
                "created_at": datetime.now(UTC),
            },
        )


class FakeSession:
    """Sessão mínima do SQLAlchemy: só o que `app.services.sessions` usa.

    Suficiente para exercitar a máquina de estados da rotação sem banco. O que
    ela não cobre — a unicidade de `token_hash`, o `rowcount` real do UPDATE —
    está na lista de validação manual do documento.
    """

    def __init__(self, stored: RefreshToken | None = None) -> None:
        self.rows: list[RefreshToken] = [stored] if stored is not None else []
        self.commits = 0
        self.mass_revocations = 0
        self.statements: list = []

    def scalar(self, statement):  # noqa: ANN001 - assinatura do SQLAlchemy
        # Reproduz o WHERE por igualdade coluna a coluna. Comparar o literal do
        # SQL seria testar o SQLAlchemy, não o serviço.
        self.statements.append(statement)
        whereclause = statement.whereclause
        filters = {
            clause.left.name: clause.right.value
            for clause in getattr(whereclause, "clauses", (whereclause,))
        }
        return next(
            (
                row
                for row in self.rows
                if all(getattr(row, name) == value for name, value in filters.items())
            ),
            None,
        )

    def add(self, row: RefreshToken) -> None:
        self.rows.append(row)

    def flush(self) -> None:
        for row in self.rows:
            if row.id is None:
                row.id = uuid.uuid4()

    def commit(self) -> None:
        self.commits += 1

    def execute(self, statement):  # noqa: ANN001 - assinatura do SQLAlchemy
        self.mass_revocations += 1
        revoked = 0
        for row in self.rows:
            if row.revoked_at is None:
                row.revoked_at = datetime.now(UTC)
                revoked += 1
        return SimpleNamespace(rowcount=revoked)


def _stored_token(token: str) -> RefreshToken:
    return RefreshToken(
        id=uuid.uuid4(),
        user_id=uuid.uuid4(),
        token_hash=hash_refresh_token(token),
        expires_at=datetime.now(UTC) + timedelta(days=30),
    )


def test_rotation_invalidates_the_presented_token() -> None:
    stored = _stored_token("token-original")
    db = FakeSession(stored)

    novo, sessao = rotate_refresh_token(db, "token-original")

    assert novo != "token-original"
    assert sessao.token_hash == hash_refresh_token(novo)
    assert stored.revoked_at is not None
    # `replaced_by_id` é o que distingue "rotacionado" de "revogado no logout".
    assert stored.replaced_by_id == sessao.id


def test_reusing_a_rotated_token_kills_every_session() -> None:
    """Reapresentar um token já rotacionado é o sinal clássico de vazamento.

    Duas partes com o mesmo segredo, e nenhuma forma de saber qual delas está
    chamando agora — a única resposta segura é derrubar tudo.
    """
    stored = _stored_token("token-vazado")
    db = FakeSession(stored)
    rotate_refresh_token(db, "token-vazado")

    with pytest.raises(RefreshTokenError, match="já utilizado"):
        rotate_refresh_token(db, "token-vazado")

    assert db.mass_revocations == 1
    assert all(row.revoked_at is not None for row in db.rows)


def test_expired_and_unknown_tokens_are_refused_without_mass_revocation() -> None:
    vencido = _stored_token("token-vencido")
    vencido.expires_at = datetime.now(UTC) - timedelta(seconds=1)
    db = FakeSession(vencido)

    with pytest.raises(RefreshTokenError, match="expirado"):
        rotate_refresh_token(db, "token-vencido")
    with pytest.raises(RefreshTokenError, match="inválido"):
        rotate_refresh_token(db, "token-que-nunca-existiu")

    # Token vencido ou desconhecido é rotina, não incidente: não pode derrubar
    # as outras sessões do usuário.
    assert db.mass_revocations == 0


def test_rotation_locks_the_row_it_is_about_to_revoke() -> None:
    """Sem `FOR UPDATE`, duas chamadas com o mesmo token rotacionam as duas.

    As duas leriam `revoked_at is None`, as duas criariam sessão nova, e a
    segunda sobrescreveria o `revoked_at`/`replaced_by_id` da primeira — dois
    refresh tokens vivos para uma rotação só, e o alarme de reuso nunca dispara.
    """
    db = FakeSession(_stored_token("token-original"))
    rotate_refresh_token(db, "token-original")

    assert db.statements[0]._for_update_arg is not None


def test_single_session_logout_works_without_knowing_the_user() -> None:
    """Quem está com o access token vencido ainda precisa conseguir deslogar."""
    stored = _stored_token("token-da-sessao")
    db = FakeSession(stored)

    assert revoke_refresh_token(db, "token-da-sessao") is True
    assert stored.revoked_at is not None
    # Mesmo motivo da rotação: ler e escrever a linha sem lock deixa dois
    # caminhos concorrentes decidirem sobre o mesmo estado.
    assert db.statements[0]._for_update_arg is not None

    # Já revogado, não há o que revogar de novo.
    assert revoke_refresh_token(db, "token-da-sessao") is False


def test_authenticated_logout_stays_restricted_to_the_caller_sessions() -> None:
    stored = _stored_token("token-de-outra-pessoa")
    db = FakeSession(stored)

    assert revoke_refresh_token(db, "token-de-outra-pessoa", uuid.uuid4()) is False
    assert stored.revoked_at is None
    assert revoke_refresh_token(db, "token-de-outra-pessoa", stored.user_id) is True


def test_logout_endpoint_accepts_a_caller_without_access_token() -> None:
    """O caso que motivou a mudança: access token vencido, sessão viva no servidor."""
    stored = _stored_token("token-da-sessao")
    db = FakeSession(stored)

    auth.logout(LogoutRequest(refresh_token="token-da-sessao"), db, None)

    assert stored.revoked_at is not None
    assert db.commits == 1


def test_logout_of_every_device_still_requires_a_valid_access_token() -> None:
    """Derrubar tudo é ação sobre a conta, não sobre o token apresentado."""
    db = FakeSession()

    with pytest.raises(HTTPException) as exc_info:
        auth.logout(LogoutRequest(all_devices=True), db, None)

    assert exc_info.value.status_code == 401
    assert db.mass_revocations == 0


def test_expired_access_token_identifies_nobody_instead_of_failing() -> None:
    """`get_optional_user` não pode dar 401: o cliente manda o header velho.

    Se a credencial vencida virasse erro, `/auth/logout` continuaria inalcançável
    justamente para quem mais precisa dele.
    """
    db = SimpleNamespace(get=lambda model, user_id: "nunca chega aqui")
    vencida = HTTPAuthorizationCredentials(scheme="Bearer", credentials="jwt.invalido")

    assert get_optional_user(db, vencida) is None
    assert get_optional_user(db, None) is None

    # Nas rotas que exigem identidade, a ausência continua sendo 401.
    with pytest.raises(HTTPException) as exc_info:
        get_current_user(None)
    assert exc_info.value.status_code == 401


# ---------------------------------------------------------------------------
# Rate limit do login (seção 7.5 do documento 07)
# ---------------------------------------------------------------------------


def _fail(state: ThrottleState, now: datetime, vezes: int) -> ThrottleState:
    for _ in range(vezes):
        state = register_failure(state, now)
    return state


def test_lockout_ladder_grows_with_each_block() -> None:
    """Cada bloqueio novo custa mais que o anterior: 10 min, 3 h, 24 h."""
    agora = datetime(2026, 8, 14, 12, tzinfo=UTC)

    quatro = _fail(ThrottleState(), agora, 4)
    # Quatro erros não bloqueiam ninguém — quem digitou errado não paga nada.
    assert quatro.lock_level == 0
    assert seconds_until_release(quatro, agora) == 0

    primeiro = _fail(ThrottleState(), agora, FAILURES_PER_LOCK)
    assert primeiro.lock_level == 1
    assert seconds_until_release(primeiro, agora) == 10 * 60

    # Cada retomada acontece assim que o bloqueio anterior termina — é o melhor
    # caso do atacante paciente, e mesmo ele sobe a escada.
    segundo = _fail(primeiro, primeiro.locked_until, FAILURES_PER_LOCK)
    assert segundo.lock_level == 2
    assert seconds_until_release(segundo, segundo.last_failure_at) == 3 * 60 * 60

    terceiro = _fail(segundo, segundo.locked_until, FAILURES_PER_LOCK)
    assert terceiro.lock_level == 3
    assert seconds_until_release(terceiro, terceiro.last_failure_at) == 24 * 60 * 60

    # Do último degrau em diante, repete 24 h — não vira bloqueio permanente.
    quarto = _fail(terceiro, terceiro.locked_until, FAILURES_PER_LOCK)
    assert quarto.lock_level == 4
    assert seconds_until_release(quarto, quarto.last_failure_at) == 24 * 60 * 60


def test_idle_time_forgets_the_count_but_not_the_step() -> None:
    """Esperar a janela não pode ser a receita para tentar em blocos de 4."""
    agora = datetime(2026, 8, 14, 12, tzinfo=UTC)
    quatro = _fail(ThrottleState(), agora, 4)

    depois = agora + FAILURE_WINDOW
    assert forget_stale_failures(quatro, depois).failures == 0

    bloqueado = _fail(ThrottleState(), agora, FAILURES_PER_LOCK)
    reincidente = forget_stale_failures(bloqueado, agora + FAILURE_WINDOW)
    assert reincidente.lock_level == 1  # o degrau conquistado permanece

    # Só o silêncio longo zera a escada inteira.
    assert forget_stale_failures(bloqueado, agora + LEVEL_DECAY) == ThrottleState()


def test_waiting_out_the_longest_block_does_not_reset_the_ladder() -> None:
    """`LEVEL_DECAY` tem de ser maior que o maior bloqueio.

    Iguais, a escada zerava no instante exato em que as 24 h terminavam: bastaria
    cumprir a punição para voltar a degraus de 10 minutos, e o teto nunca valeria.
    """
    assert LEVEL_DECAY > lock_duration(len(LOCK_DURATIONS))


def test_ip_scope_tolerates_more_failures_than_the_email_scope() -> None:
    """Um IP pode ser um escritório inteiro atrás de NAT."""
    agora = datetime(2026, 8, 14, 12, tzinfo=UTC)
    state = ThrottleState()
    for passo in range(FAILURES_PER_LOCK):
        state = register_failure(
            state,
            agora + timedelta(seconds=passo),
            failures_per_lock=FAILURES_PER_LOCK * IP_TOLERANCE,
        )
    assert state.lock_level == 0
    assert seconds_until_release(state, agora) == 0


def test_throttle_never_stores_the_email_or_the_ip() -> None:
    """Um dump da tabela não pode virar lista de contas cadastradas."""
    hash_ = scope_hash("email", "pessoa@example.com")

    assert "pessoa@example.com" not in hash_
    assert len(hash_) == 64  # SHA-256 em hexadecimal cabe no String(64).
    # Escopos diferentes não colidem mesmo com o mesmo valor.
    assert scope_hash("ip", "pessoa@example.com") != hash_


def test_blocked_login_answers_429_without_touching_the_password() -> None:
    """Bloqueio existe para não fazer o trabalho — nem bcrypt, nem consulta."""

    class RecusaTudo:
        def scalar(self, statement):  # noqa: ANN001 - assinatura do SQLAlchemy
            raise AssertionError("login bloqueado não pode consultar o usuário")

    def nao_chamar(*args, **kwargs):  # noqa: ANN002, ANN003
        raise AssertionError("login bloqueado não pode verificar senha")

    with pytest.MonkeyPatch.context() as patch:
        patch.setattr(auth, "verify_password", nao_chamar)
        patch.setattr(
            auth,
            "check_login_allowed",
            lambda db, email, ip: (_ for _ in ()).throw(LoginBlocked(600)),
        )
        with pytest.raises(HTTPException) as exc_info:
            auth.login(
                LoginRequest(email="alvo@example.com", password="tentativa"),
                RecusaTudo(),
            )

    assert exc_info.value.status_code == 429
    assert exc_info.value.headers["Retry-After"] == "600"


def test_change_password_defaults_to_revoking_other_devices() -> None:
    payload = ChangePasswordRequest(
        current_password="senha-antiga", new_password="senha-nova-seguraaa"
    )
    assert payload.revoke_other_sessions is True

    with pytest.raises(ValidationError, match="72 bytes"):
        ChangePasswordRequest(
            current_password="senha-antiga", new_password="á" * 37
        )


def test_profile_update_only_exposes_the_name() -> None:
    """Trocar e-mail muda a identidade de login e ficou fora deste endpoint."""
    assert set(ProfileUpdate.model_fields) == {"full_name"}


# ---------------------------------------------------------------------------
# B1 — edição de movimentação
# ---------------------------------------------------------------------------


def _existing_purchase() -> dict:
    return {
        "movement_type": "purchase",
        "occurred_at": datetime(2026, 8, 1, 12, tzinfo=UTC),
        "quantity": Decimal("100"),
        "unit_price": Decimal("30.50"),
        "costs": Decimal("0"),
        "gross_amount": None,
        "net_amount": None,
        "factor": None,
        "trade_kind": "swing_trade",
        "fx_rate": None,
        "fx_rate_date": None,
        "notes": None,
    }


def test_movement_update_accepts_a_partial_payload() -> None:
    """Sozinho o PATCH não valida regra por tipo — só o resultado da fusão vale."""
    payload = MovementUpdate(unit_price="31.20")
    assert payload.model_dump(exclude_unset=True) == {"unit_price": Decimal("31.20")}


def test_merged_movement_is_validated_as_a_whole() -> None:
    merged = _existing_purchase()
    merged.update(MovementUpdate(quantity="150").model_dump(exclude_unset=True))
    validated = MovementCreate.model_validate(merged)
    assert validated.quantity == Decimal("150")

    # Virar provento sem valor bruto nem líquido tem de falhar, mesmo que o corpo
    # do PATCH só tenha trocado o tipo: a regra depende do estado final.
    virou_provento = _existing_purchase()
    virou_provento.update(
        MovementUpdate(movement_type="dividend").model_dump(exclude_unset=True)
    )
    virou_provento["quantity"] = None
    virou_provento["unit_price"] = None
    with pytest.raises(ValidationError, match="provento"):
        MovementCreate.model_validate(virou_provento)


# ---------------------------------------------------------------------------
# B2 — exportação da carteira
# ---------------------------------------------------------------------------


def _position(ticker: str = "PETR4", *, with_quote: bool = True) -> SimpleNamespace:
    quote = (
        SimpleNamespace(price=Decimal("31.20"), collected_at=datetime(2026, 8, 12, tzinfo=UTC))
        if with_quote
        else None
    )
    return SimpleNamespace(
        asset=SimpleNamespace(
            ticker=ticker, name="=Petrobras", asset_type="stock", currency="BRL"
        ),
        quantity=Decimal("100"),
        average_price=Decimal("30.50"),
        invested_cost=Decimal("3050.00"),
        realized_gain=Decimal("0"),
        dividends_gross=Decimal("0"),
        dividends_net=Decimal("0"),
        quote=quote,
        market_value=Decimal("3120.00") if with_quote else None,
        unrealized_gain=Decimal("70.00") if with_quote else None,
        return_on_cost=Decimal("0.022951") if with_quote else None,
    )


def test_position_export_neutralizes_formulas_and_keeps_nulls_empty() -> None:
    content = export_positions_csv([_position(with_quote=False)]).decode("utf-8-sig")
    header, row = content.splitlines()[:2]
    cells = row.split(";")
    columns = dict(zip(header.split(";"), cells, strict=True))

    # Nome começando com "=" viraria fórmula ao abrir no Excel.
    assert columns["nome"] == "'=Petrobras"
    # Sem cotação, valor de mercado é vazio — não zero. Zero diria que a posição
    # não vale nada, que é uma afirmação diferente de "não sei quanto vale".
    assert columns["valor_mercado"] == ""
    assert columns["cotacao"] == ""
    assert columns["quantidade"] == "100"


def test_portfolio_xlsx_has_one_sheet_per_view() -> None:
    movement = SimpleNamespace(
        occurred_at=datetime(2026, 8, 1, 12, tzinfo=UTC),
        movement_type="purchase",
        quantity=Decimal("100"),
        unit_price=Decimal("30.50"),
        costs=Decimal("0"),
        gross_amount=None,
        net_amount=None,
        factor=None,
        trade_kind="swing_trade",
        fx_rate=None,
        fx_rate_date=None,
        notes=None,
    )
    content = export_portfolio_xlsx([_position()], [(movement, "PETR4")])
    workbook = load_workbook(io.BytesIO(content), read_only=True)
    try:
        assert workbook.sheetnames == ["Posições", "Movimentações"]
    finally:
        workbook.close()


# ---------------------------------------------------------------------------
# Registro das rotas novas
# ---------------------------------------------------------------------------


def test_gap_endpoints_are_registered() -> None:
    paths = app.openapi()["paths"]
    assert {
        "/api/v1/auth/refresh",
        "/api/v1/auth/logout",
        "/api/v1/auth/change-password",
        "/api/v1/dashboard/timeseries",
        "/api/v1/transactions/categories",
        "/api/v1/transactions/imports",
        "/api/v1/investments/snapshots",
        "/api/v1/investments/export",
        "/api/v1/investments/movements/{movement_id}",
    } <= set(paths)
    assert "patch" in paths["/api/v1/auth/me"]
    assert "delete" in paths["/api/v1/telegram/link"]
