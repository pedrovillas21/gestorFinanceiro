from csv import reader
from datetime import UTC, datetime, timedelta
from decimal import Decimal
from io import BytesIO, StringIO
from types import SimpleNamespace
from zoneinfo import ZoneInfo

import pytest
from openpyxl import load_workbook

from app.core.config import settings
from app.services.spreadsheets import (
    _format_brl,
    _format_period_label,
    export_csv,
    export_transactions_pdf,
    export_xlsx,
)
from app.services.telegram_bot import (
    _buscar_pendencia,
    _consentimento_valido,
    criar_link_token,
)


class PendingDb:
    def __init__(self, pending):
        self.pending = pending
        self.deleted = []
        self.commits = 0

    def scalar(self, statement):
        return self.pending

    def delete(self, value) -> None:
        self.deleted.append(value)

    def commit(self) -> None:
        self.commits += 1


def test_expired_pending_transaction_is_deleted_and_not_returned() -> None:
    pending = SimpleNamespace(expires_at=datetime.now(UTC) - timedelta(seconds=1))
    db = PendingDb(pending)

    assert _buscar_pendencia(db, "123") is None
    assert db.deleted == [pending]
    assert db.commits == 1


def test_active_pending_transaction_is_returned() -> None:
    pending = SimpleNamespace(expires_at=datetime.now(UTC) + timedelta(minutes=1))
    db = PendingDb(pending)

    assert _buscar_pendencia(db, "123") is pending
    assert db.deleted == []
    assert db.commits == 0


def test_only_the_current_published_policy_is_valid_consent() -> None:
    current = SimpleNamespace(
        privacy_consented_at=datetime.now(UTC),
        privacy_consent_version=settings.PRIVACY_POLICY_VERSION,
    )
    stale = SimpleNamespace(
        privacy_consented_at=datetime.now(UTC),
        privacy_consent_version="versao-antiga",
    )

    assert _consentimento_valido(current)
    assert not _consentimento_valido(stale)


def test_link_generation_rejects_unpublished_policy_before_database_access(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class UntouchedDb:
        def scalars(self, statement):
            raise AssertionError("o banco não deve ser acessado")

    def unpublished_policy(version: str) -> None:
        assert version == settings.PRIVACY_POLICY_VERSION
        raise KeyError(version)

    monkeypatch.setattr(
        "app.services.telegram_bot.get_privacy_policy", unpublished_policy
    )

    with pytest.raises(ValueError, match="não foi publicada"):
        criar_link_token(
            UntouchedDb(),
            SimpleNamespace(),
            consent_version=settings.PRIVACY_POLICY_VERSION,
        )


def _formula_transaction():
    return SimpleNamespace(
        occurred_at=datetime(2026, 8, 10, tzinfo=UTC),
        type="expense",
        description="=HYPERLINK(\"https://example.invalid\")",
        amount=Decimal("10.00"),
        category="+cmd|' /C calc'!A0",
        payment_method="@SUM(1+1)",
    )


def test_csv_export_escapes_formula_prefixes() -> None:
    content = export_csv([_formula_transaction()]).decode("utf-8-sig")
    row = list(reader(StringIO(content), delimiter=";"))[1]

    assert row[2].startswith("'=")
    assert row[4].startswith("'+")
    assert row[5].startswith("'@")


def test_xlsx_export_escapes_formula_prefixes() -> None:
    workbook = load_workbook(BytesIO(export_xlsx([_formula_transaction()])))
    try:
        row = next(workbook.active.iter_rows(min_row=2, max_row=2))
        assert row[2].value.startswith("'=")
        assert row[4].value.startswith("'+")
        assert row[5].value.startswith("'@")
        assert all(row[index].data_type == "s" for index in (2, 4, 5))
    finally:
        workbook.close()


def test_format_brl_groups_thousands_from_decimal_without_float() -> None:
    assert _format_brl(Decimal("1234567.9")) == "R$ 1.234.567,90"
    assert _format_brl(Decimal("0.5")) == "R$ 0,50"
    assert _format_brl(Decimal("-42.90")) == "-R$ 42,90"


def test_format_period_label_shows_last_included_day_for_exclusive_end() -> None:
    # `end` é exclusivo (backend/app/api/v1/transactions.py `_conditions`): o
    # rótulo do relatório precisa mostrar 31/08, não 01/09. Construído com o
    # offset de America/Sao_Paulo (não UTC) porque é assim que o front sempre
    # emite essas datas (lib/format.ts `toIsoWithOffset`) — meia-noite UTC
    # legitimamente vira 31/07 à noite em SP, então testar com UTC aqui
    # denunciaria um bug que não existe.
    tz = ZoneInfo("America/Sao_Paulo")
    start = datetime(2026, 8, 1, tzinfo=tz)
    end = datetime(2026, 9, 1, tzinfo=tz)
    assert _format_period_label(start, end) == "Período: 01/08/2026 até 31/08/2026"
    assert _format_period_label(None, None) == "Período: todo o histórico"
    assert _format_period_label(start, None) == "Período: 01/08/2026 até hoje"


def test_pdf_export_produces_a_valid_document_and_escapes_markup() -> None:
    """Descrição com `&`/`<` não pode quebrar o Paragraph do reportlab (que interpreta um subconjunto de XML)."""
    tricky = SimpleNamespace(
        occurred_at=datetime(2026, 8, 10, tzinfo=UTC),
        type="income",
        description="Salário <julho> & bônus",
        amount=Decimal("1234.56"),
        category=None,
        payment_method=None,
    )
    content = export_transactions_pdf(
        [tricky, _formula_transaction()],
        start=datetime(2026, 8, 1, tzinfo=UTC),
        end=datetime(2026, 9, 1, tzinfo=UTC),
    )
    assert content.startswith(b"%PDF")
    assert len(content) > 0


def test_pdf_export_handles_an_empty_period() -> None:
    content = export_transactions_pdf([], start=None, end=None)
    assert content.startswith(b"%PDF")
