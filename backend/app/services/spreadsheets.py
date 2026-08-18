"""Importação e exportação de transações sem executar pandas no event loop."""

import asyncio
import csv
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal
from io import BytesIO, StringIO
import logging
from pathlib import Path
import uuid
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from sqlalchemy import and_, or_, select
from sqlalchemy.orm import Session, undefer

from app.core.money import quantize_money
from app.database import SessionLocal
from app.models.import_job import ImportJob
from app.models.transaction import Transaction


TZ = ZoneInfo("America/Sao_Paulo")
MAX_IMPORT_ROWS = 10_000
IMPORT_JOB_LEASE = timedelta(minutes=5)
IMPORT_WORKER_POLL_SECONDS = 1.0
logger = logging.getLogger(__name__)
COLUMN_ALIASES = {
    "description": ("description", "descricao", "descrição"),
    "amount": ("amount", "valor"),
    "category": ("category", "categoria"),
    "type": ("type", "tipo"),
    "payment_method": ("payment_method", "metodo_pagamento", "método_pagamento"),
    "occurred_at": ("occurred_at", "data", "date"),
}
TYPE_ALIASES = {
    "income": "income",
    "receita": "income",
    "entrada": "income",
    "expense": "expense",
    "despesa": "expense",
    "saida": "expense",
    "saída": "expense",
}


def _normalized_headers(row: dict[object, object]) -> dict[str, object]:
    raw = {str(key).strip().lower(): value for key, value in row.items() if key is not None}
    return {
        canonical: next((raw[name] for name in aliases if name in raw), None)
        for canonical, aliases in COLUMN_ALIASES.items()
    }


def _parse_datetime(value: object) -> datetime:
    if value in (None, ""):
        return datetime.now(UTC)
    if isinstance(value, datetime):
        result = value
    elif isinstance(value, date):
        result = datetime.combine(value, datetime.min.time())
    else:
        text = str(value).strip()
        try:
            result = datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError as exc:
            raise ValueError(f"data inválida: {text}; use AAAA-MM-DD") from exc
    if result.tzinfo is None:
        result = result.replace(tzinfo=TZ)
    return result.astimezone(UTC)


def _limited_rows(rows, *, max_rows: int = MAX_IMPORT_ROWS) -> list:
    limited = []
    for row_number, row in enumerate(rows, start=1):
        if row_number > max_rows:
            raise ValueError(f"arquivo excede o limite de {max_rows} linhas")
        limited.append(row)
    return limited


def _csv_rows(
    content: bytes, *, max_rows: int = MAX_IMPORT_ROWS
) -> list[dict[object, object]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;")
    except csv.Error:
        dialect = csv.excel
    return _limited_rows(
        csv.DictReader(StringIO(text), dialect=dialect), max_rows=max_rows
    )


def _xlsx_rows(
    content: bytes, *, max_rows: int = MAX_IMPORT_ROWS
) -> list[dict[object, object]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return []
        return _limited_rows(
            (dict(zip(headers, values, strict=False)) for values in rows),
            max_rows=max_rows,
        )
    finally:
        workbook.close()


def parse_rows(
    content: bytes, filename: str, *, max_rows: int = MAX_IMPORT_ROWS
) -> list[dict[str, object]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        source_rows = _csv_rows(content, max_rows=max_rows)
    elif suffix == ".xlsx":
        source_rows = _xlsx_rows(content, max_rows=max_rows)
    else:
        raise ValueError("formato não suportado; envie CSV ou XLSX")

    parsed: list[dict[str, object]] = []
    for row_number, raw_row in enumerate(source_rows, start=2):
        row = _normalized_headers(raw_row)
        description = str(row["description"] or "").strip()
        transaction_type = TYPE_ALIASES.get(str(row["type"] or "").strip().lower())
        if not description or transaction_type is None or row["amount"] in (None, ""):
            raise ValueError(
                f"linha {row_number}: descrição, valor e tipo (receita/despesa) são obrigatórios"
            )
        amount = quantize_money(row["amount"])
        if amount <= 0:
            raise ValueError(f"linha {row_number}: valor deve ser positivo")
        parsed.append(
            {
                "description": description[:255],
                "amount": amount,
                "category": str(row["category"]).strip()[:100]
                if row["category"] not in (None, "")
                else None,
                "type": transaction_type,
                "payment_method": str(row["payment_method"]).strip()[:50]
                if row["payment_method"] not in (None, "")
                else None,
                "occurred_at": _parse_datetime(row["occurred_at"]),
            }
        )
    return parsed


def import_transactions(
    db: Session,
    user_id: uuid.UUID,
    content: bytes,
    filename: str,
    *,
    commit: bool = True,
) -> tuple[int, int]:
    rows = parse_rows(content, filename)
    for row in rows:
        db.add(Transaction(user_id=user_id, source="import", **row))
    if commit:
        db.commit()
    return len(rows), len(rows)


def _recoverable_job_filter(now: datetime):
    stale_before = now - IMPORT_JOB_LEASE
    return or_(
        ImportJob.status == "pending",
        and_(
            ImportJob.status == "processing",
            or_(
                ImportJob.processing_started_at.is_(None),
                ImportJob.processing_started_at < stale_before,
            ),
        ),
    )


def process_import_job(job_id: uuid.UUID, user_id: uuid.UUID) -> bool:
    """Claims and processes one durable import job.

    The imported transactions and terminal job state share one commit, so a
    process crash cannot leave committed rows behind for a later retry to copy.
    """
    db = SessionLocal()
    try:
        now = datetime.now(UTC)
        job = db.scalar(
            select(ImportJob)
            .options(undefer(ImportJob.content))
            .where(
                ImportJob.id == job_id,
                ImportJob.user_id == user_id,
                _recoverable_job_filter(now),
            )
            .with_for_update(skip_locked=True)
        )
        if job is None:
            return False
        content = job.content
        filename = job.filename
        if content is None:
            job.status = "failed"
            job.error_message = "Conteúdo da importação não está disponível"
            job.completed_at = now
            db.commit()
            return True
        job.status = "processing"
        job.processing_started_at = now
        job.attempt_count += 1
        db.commit()
        try:
            rows = parse_rows(content, filename)
            job = db.get(ImportJob, job_id)
            if job is not None:
                for row in rows:
                    db.add(Transaction(user_id=user_id, source="import", **row))
                job.status = "completed"
                job.total_rows = len(rows)
                job.imported_rows = len(rows)
                job.completed_at = datetime.now(UTC)
                job.content = None
                db.commit()
        except Exception as exc:  # o estado do job precisa registrar qualquer falha
            db.rollback()
            job = db.get(ImportJob, job_id)
            if job is not None:
                job.status = "failed"
                job.error_message = str(exc)[:2000]
                job.completed_at = datetime.now(UTC)
                job.content = None
                db.commit()
        return True
    finally:
        db.close()


def process_next_import_job() -> bool:
    db = SessionLocal()
    try:
        now = datetime.now(UTC)
        candidate = db.execute(
            select(ImportJob.id, ImportJob.user_id)
            .where(_recoverable_job_filter(now))
            .order_by(ImportJob.created_at)
            .limit(1)
        ).first()
    finally:
        db.close()
    if candidate is None:
        return False
    return process_import_job(candidate.id, candidate.user_id)


async def run_import_worker(stop_event: asyncio.Event) -> None:
    """Continuously recovers pending jobs and jobs abandoned after a crash."""
    while not stop_event.is_set():
        try:
            processed = await asyncio.to_thread(process_next_import_job)
        except Exception:
            logger.exception("Falha ao consultar a fila durável de importações")
            processed = False
        if processed:
            continue
        try:
            await asyncio.wait_for(
                stop_event.wait(), timeout=IMPORT_WORKER_POLL_SECONDS
            )
        except TimeoutError:
            pass


def _escape_spreadsheet_formula(value: str | None) -> str:
    """Mantém texto controlado pelo usuário inerte em CSV e XLSX."""
    if not value:
        return ""
    if value.startswith(("=", "+", "-", "@", "\t", "\r", "\n")):
        return f"'{value}"
    return value


def export_csv(transactions: list[Transaction]) -> bytes:
    output = StringIO(newline="")
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["data", "tipo", "descricao", "valor", "categoria", "metodo_pagamento"])
    for item in transactions:
        writer.writerow(
            [
                item.occurred_at.isoformat(),
                item.type,
                _escape_spreadsheet_formula(item.description),
                str(item.amount),
                _escape_spreadsheet_formula(item.category),
                _escape_spreadsheet_formula(item.payment_method),
            ]
        )
    return output.getvalue().encode("utf-8-sig")


def export_xlsx(transactions: list[Transaction]) -> bytes:
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Transações")
    sheet.append(["data", "tipo", "descricao", "valor", "categoria", "metodo_pagamento"])
    for item in transactions:
        sheet.append(
            [
                item.occurred_at.replace(tzinfo=None),
                item.type,
                _escape_spreadsheet_formula(item.description),
                str(item.amount),
                _escape_spreadsheet_formula(item.category),
                _escape_spreadsheet_formula(item.payment_method),
            ]
        )
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


# --------------------------------------------------------------------------------------
# Exportação de transações em PDF — relatório legível/imprimível. O CSV cobre
# reimportação em outra planilha; este cobre leitura e compartilhamento
# (extrato formatado, com totais), então os dois deixam de ser redundantes.
# --------------------------------------------------------------------------------------

_PDF_TYPE_LABEL = {"income": "Receita", "expense": "Despesa"}
_PDF_SUCCESS_COLOR = "#16a34a"  # mesmo token --success de frontend/app/globals.css (tema claro)
_PDF_DANGER_COLOR = "#dc2626"  # mesmo token --danger


def _format_brl(value: Decimal) -> str:
    """Mesma regra dura do resto do projeto: formata a partir do Decimal armazenado, nunca de um float intermediário."""
    quantized = value.quantize(Decimal("0.01"))
    sign = "-" if quantized < 0 else ""
    integer_part, _, fraction_part = f"{abs(quantized):f}".partition(".")
    groups: list[str] = []
    while len(integer_part) > 3:
        groups.insert(0, integer_part[-3:])
        integer_part = integer_part[:-3]
    groups.insert(0, integer_part)
    return f"{sign}R$ {'.'.join(groups)},{fraction_part}"


def _format_datetime_ptbr(value: datetime) -> str:
    return value.astimezone(TZ).strftime("%d/%m/%Y %H:%M")


def _format_date_ptbr(value: datetime) -> str:
    return value.astimezone(TZ).strftime("%d/%m/%Y")


def _format_period_label(start: datetime | None, end: datetime | None) -> str:
    """`end` é exclusivo (contrato de `_conditions`) — o rótulo mostra o último dia incluído, não `end` em si."""
    if start is None and end is None:
        return "Período: todo o histórico"
    start_label = _format_date_ptbr(start) if start is not None else "o início"
    end_label = _format_date_ptbr(end - timedelta(microseconds=1)) if end is not None else "hoje"
    return f"Período: {start_label} até {end_label}"


def export_transactions_pdf(
    transactions: list[Transaction], start: datetime | None, end: datetime | None
) -> bytes:
    from xml.sax.saxutils import escape

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    styles = getSampleStyleSheet()
    cell_style = styles["Normal"].clone("cell")
    cell_style.fontSize = 8
    cell_style.leading = 10

    def cell(text: str) -> Paragraph:
        return Paragraph(escape(text), cell_style)

    story = [
        Paragraph("Extrato de transações", styles["Title"]),
        Paragraph(_format_period_label(start, end), styles["Normal"]),
        Paragraph(f"Gerado em {_format_datetime_ptbr(datetime.now(UTC))}", styles["Normal"]),
        Spacer(1, 0.5 * cm),
    ]

    header = ["Data", "Tipo", "Descrição", "Categoria", "Forma de pagamento", "Valor"]
    rows: list[list[object]] = [header]
    income_row_indexes: list[int] = []
    expense_row_indexes: list[int] = []
    total_income = Decimal("0")
    total_expense = Decimal("0")
    for item in transactions:
        is_income = item.type == "income"
        (income_row_indexes if is_income else expense_row_indexes).append(len(rows))
        if is_income:
            total_income += item.amount
        else:
            total_expense += item.amount
        rows.append(
            [
                _format_datetime_ptbr(item.occurred_at),
                _PDF_TYPE_LABEL.get(item.type, item.type),
                cell(item.description),
                cell(item.category or "Sem categoria"),
                cell(item.payment_method or "—"),
                _format_brl(item.amount),
            ]
        )

    # Larguras somam 17,4cm — cabem nos 18cm úteis do A4 (21cm - 2x1,5cm de margem).
    table = Table(
        rows,
        colWidths=[2.6 * cm, 1.6 * cm, 5.0 * cm, 3.0 * cm, 2.8 * cm, 2.4 * cm],
        repeatRows=1,
    )
    style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f8f8f8")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#171717")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (1, -1), "CENTER"),
        ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e4e4e7")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f8f8")]),
    ]
    for row_index in income_row_indexes:
        style_commands.append(("TEXTCOLOR", (-1, row_index), (-1, row_index), colors.HexColor(_PDF_SUCCESS_COLOR)))
    for row_index in expense_row_indexes:
        style_commands.append(("TEXTCOLOR", (-1, row_index), (-1, row_index), colors.HexColor(_PDF_DANGER_COLOR)))
    table.setStyle(TableStyle(style_commands))
    story.append(table)

    balance = total_income - total_expense
    story.append(Spacer(1, 0.5 * cm))
    story.append(Paragraph(f"Total de receitas: {_format_brl(total_income)}", styles["Normal"]))
    story.append(Paragraph(f"Total de despesas: {_format_brl(total_expense)}", styles["Normal"]))
    story.append(Paragraph(f"<b>Saldo: {_format_brl(balance)}</b>", styles["Normal"]))

    buffer = BytesIO()
    SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        title="Extrato de transações",
    ).build(story)
    return buffer.getvalue()


# --------------------------------------------------------------------------------------
# Exportação da carteira de investimentos
# --------------------------------------------------------------------------------------

POSITION_HEADERS = [
    "ticker",
    "nome",
    "tipo",
    "moeda",
    "quantidade",
    "preco_medio",
    "custo_investido",
    "cotacao",
    "cotacao_coletada_em",
    "valor_mercado",
    "ganho_nao_realizado",
    "ganho_realizado",
    "proventos_bruto",
    "proventos_liquido",
    "retorno_sobre_custo",
]
MOVEMENT_HEADERS = [
    "data",
    "ticker",
    "tipo_movimento",
    "quantidade",
    "preco_unitario",
    "custos",
    "valor_bruto",
    "valor_liquido",
    "fator",
    "operacao",
    "cambio",
    "cambio_data",
    "observacoes",
]


def _optional(value: object) -> str:
    """Nulo vira célula vazia, nunca "None" nem 0 — a distinção importa aqui.

    Uma posição sem cotação tem `market_value` nulo; escrever 0 diria que a
    posição vale zero, que é outra coisa.
    """
    return "" if value is None else str(value)


def _position_row(position) -> list[str]:
    quote = position.quote
    return [
        _escape_spreadsheet_formula(position.asset.ticker),
        _escape_spreadsheet_formula(position.asset.name),
        _escape_spreadsheet_formula(position.asset.asset_type),
        _escape_spreadsheet_formula(position.asset.currency),
        _optional(position.quantity),
        _optional(position.average_price),
        _optional(position.invested_cost),
        _optional(quote.price if quote else None),
        _optional(quote.collected_at.isoformat() if quote else None),
        _optional(position.market_value),
        _optional(position.unrealized_gain),
        _optional(position.realized_gain),
        _optional(position.dividends_gross),
        _optional(position.dividends_net),
        _optional(position.return_on_cost),
    ]


def _movement_row(movement, ticker: str) -> list[str]:
    return [
        movement.occurred_at.isoformat(),
        _escape_spreadsheet_formula(ticker),
        _escape_spreadsheet_formula(movement.movement_type),
        _optional(movement.quantity),
        _optional(movement.unit_price),
        _optional(movement.costs),
        _optional(movement.gross_amount),
        _optional(movement.net_amount),
        _optional(movement.factor),
        _escape_spreadsheet_formula(movement.trade_kind),
        _optional(movement.fx_rate),
        _optional(movement.fx_rate_date.isoformat() if movement.fx_rate_date else None),
        _escape_spreadsheet_formula(movement.notes),
    ]


def _rows_to_csv(headers: list[str], rows: list[list[str]]) -> bytes:
    output = StringIO(newline="")
    writer = csv.writer(output, delimiter=";")
    writer.writerow(headers)
    writer.writerows(rows)
    return output.getvalue().encode("utf-8-sig")


def export_positions_csv(positions: list) -> bytes:
    return _rows_to_csv(POSITION_HEADERS, [_position_row(item) for item in positions])


def export_movements_csv(movements: list[tuple[object, str]]) -> bytes:
    """`movements` são pares (movimentação, ticker) — o ticker não está na linha."""
    return _rows_to_csv(
        MOVEMENT_HEADERS, [_movement_row(item, ticker) for item, ticker in movements]
    )


def export_portfolio_xlsx(
    positions: list, movements: list[tuple[object, str]]
) -> bytes:
    """Carteira em duas abas: consolidado e extrato.

    Cada uma responde a uma pergunta diferente — "quanto tenho hoje" e "como
    cheguei aqui" —, e o XLSX é o único formato dos dois que comporta as duas
    sem obrigar o usuário a baixar dois arquivos.
    """
    workbook = Workbook(write_only=True)
    positions_sheet = workbook.create_sheet("Posições")
    positions_sheet.append(POSITION_HEADERS)
    for position in positions:
        positions_sheet.append(_position_row(position))

    movements_sheet = workbook.create_sheet("Movimentações")
    movements_sheet.append(MOVEMENT_HEADERS)
    for movement, ticker in movements:
        movements_sheet.append(_movement_row(movement, ticker))

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
